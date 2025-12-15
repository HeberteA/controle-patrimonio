import streamlit as st
import pandas as pd
from st_supabase_connection import SupabaseConnection
from streamlit_option_menu import option_menu 
import base64
import io
import time
import tempfile  
from datetime import datetime
import plotly.express as px
from fpdf import FPDF 
import openpyxl 
import qrcode 
from PIL import Image  

st.set_page_config(
    page_title="Controle de Patrimônio Lavie",
    page_icon="Lavie1.png",
    layout="wide"
)

APP_STYLE_CSS = """
<style>
[data-testid="stAppViewContainer"] {
    background: radial-gradient(circle at 10% 20%, #1e1e24 0%, #050505 90%);
    background-attachment: fixed;
}
/* Ajustes de Inputs para contraste */
div[data-baseweb="input"] > div, div[data-baseweb="select"] > div {
    background-color: rgba(255, 255, 255, 0.05) !important;
    border: 1px solid rgba(255, 255, 255, 0.1) !important;
    color: white !important;
}
div[data-testid="stNumberInput"] input, div[data-testid="stTextInput"] input {
    color: white !important;
}
/* Steps */
.step-container {
    display: flex; justify-content: space-between; align-items: center; margin-bottom: 40px;
    background: rgba(255,255,255,0.03); padding: 20px; border-radius: 50px; border: 1px solid rgba(255,255,255,0.05);
}
.step-item {
    display: flex; align-items: center; flex-direction: column; color: #666; font-weight: 500; width: 33%; position: relative;
}
.step-item .step-number {
    width: 35px; height: 35px; border-radius: 50%; border: 2px solid #555; display: flex; align-items: center; justify-content: center;
    font-weight: bold; margin-bottom: 8px; transition: all 0.3s ease; background-color: #111;
}
.step-item.active { color: #E37026; }
.step-item.active .step-number { border-color: #E37026; background-color: #E37026; color: #FFFFFF; box-shadow: 0 0 15px rgba(227, 112, 38, 0.5); }
</style>
"""
st.markdown(APP_STYLE_CSS, unsafe_allow_html=True)

if 'logged_in' not in st.session_state:
    st.session_state.logged_in = False
if 'is_admin' not in st.session_state:
    st.session_state.is_admin = False
if 'selected_obra' not in st.session_state:
    st.session_state.selected_obra = None
if 'edit_item_id' not in st.session_state:
    st.session_state.edit_item_id = None 
if 'confirm_delete' not in st.session_state:
    st.session_state.confirm_delete = False
if 'movement_item_id' not in st.session_state:
    st.session_state.movement_item_id = None

ID_COL = "id"
OBRA_COL = "obra"
TOMBAMENTO_COL = "numero_tombamento"
NOME_COL = "nome"
STATUS_COL = "status"
NF_NUM_COL = "numero_nota_fiscal"
NF_LINK_COL = "link_nota_fiscal"
ESPEC_COL = "especificacoes"
OBS_COL = "observacoes"
LOCAL_COL = "local_de_uso"
RESPONSAVEL_COL = "responsavel"
VALOR_COL = "valor"

def get_img_as_base64(file):
    try:
        with open(file, "rb") as f:
            data = f.read()
        return base64.b64encode(data).decode()
    except Exception:
        return None

def upload_to_supabase_storage(file_data, file_name, file_type='application/pdf'):
    try:
        conn_storage = st.connection(
            "supabase",
            type=SupabaseConnection,
            url=st.secrets["connections"]["supabase"]["url"],
            key=st.secrets["connections"]["supabase"]["key"]
        )
        bucket_name = "notas-fiscais"
        
        conn_storage.storage.from_(bucket_name).upload(
            file=file_data,
            path=file_name,
            file_options={"content-type": file_type, "x-upsert": "true"}
        )
        
        response = conn_storage.storage.from_(bucket_name).get_public_url(file_name)
        return response
    
    except Exception as e:
        st.error(f"Erro no upload para o Supabase Storage: {e}")
        return None

def gerar_numero_tombamento_sequencial(existing_data, obra_para_gerar):
    if not obra_para_gerar: return None
    itens = existing_data[existing_data[OBRA_COL] == obra_para_gerar]
    if itens.empty: return "1"
    numeros_numericos = pd.to_numeric(itens[TOMBAMENTO_COL], errors='coerce').dropna()
    if numeros_numericos.empty: return "1"
    return str(int(numeros_numericos.max()) + 1)

try:
    conn = st.connection(
        "supabase",
        type=SupabaseConnection,
        url=st.secrets["connections"]["supabase"]["url"],
        key=st.secrets["connections"]["supabase"]["key"]
    )
except Exception as e:
    st.error("ERRO GRAVE NA CONEXÃO COM O SUPABASE. Verifique os secrets.")
    st.exception(e)
    st.stop()


@st.cache_data(ttl=30) 
def carregar_dados_app():
    try:
        status_resp = conn.table("status").select("*").execute()
        lista_status = [row['nome_do_status'] for row in status_resp.data]
        
        obras_resp = conn.table("obras").select("*").execute()
        lista_obras = [row['nome_da_obra'] for row in obras_resp.data]
        
        patrimonio_resp = conn.table("patrimonio").select("*").execute()
        patrimonio_df = pd.DataFrame(patrimonio_resp.data)
        
        colunas_patrimonio = [
            ID_COL, OBRA_COL, TOMBAMENTO_COL, NOME_COL, ESPEC_COL, 
            OBS_COL, LOCAL_COL, RESPONSAVEL_COL, NF_NUM_COL, 
            NF_LINK_COL, VALOR_COL, STATUS_COL
        ]

        if patrimonio_df.empty: 
             patrimonio_df = pd.DataFrame(columns=colunas_patrimonio)
        
        if VALOR_COL in patrimonio_df.columns:
            patrimonio_df[VALOR_COL] = pd.to_numeric(patrimonio_df[VALOR_COL], errors='coerce').fillna(0)
        
        movimentacoes_resp = conn.table("movimentacoes").select("*").execute()
        movimentacoes_df = pd.DataFrame(movimentacoes_resp.data)
        if movimentacoes_df.empty:
            movimentacoes_df = pd.DataFrame(columns=[
                ID_COL, OBRA_COL, TOMBAMENTO_COL, "tipo_movimentacao", 
                "data_hora", "responsavel_movimentacao", "observacoes"
            ])
            
        locacoes_resp = conn.table("locacoes").select("*").execute()
        locacoes_df = pd.DataFrame(locacoes_resp.data)
        colunas_locacao = [
            "id", "equipamento", "obra_destino", "responsavel", "quantidade", 
            "unidade", "valor_mensal", "contrato_sienge", "status", 
            "data_inicio", "data_previsao_fim"
        ]
        
        if locacoes_df.empty:
            locacoes_df = pd.DataFrame(columns=colunas_locacao)
        else:
            locacoes_df['data_inicio'] = pd.to_datetime(locacoes_df['data_inicio'], errors='coerce')
            locacoes_df['data_previsao_fim'] = pd.to_datetime(locacoes_df['data_previsao_fim'], errors='coerce')

        return lista_status, lista_obras, patrimonio_df, movimentacoes_df, locacoes_df
    
    except KeyError as e:
        st.error(f"Erro Crítico de 'KeyError'. Coluna não encontrada: {e}")
        return [], [], pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    except Exception as e:
        st.error(f"Erro ao carregar dados do Supabase: {e}")
        return [], [], pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

@st.cache_data
def to_excel(df):
    """Converte DataFrame para um arquivo Excel em memória."""
    try:
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Patrimonio')
        processed_data = output.getvalue()
        return processed_data
    except Exception as e:
        st.error(f"Erro ao gerar Excel: {e}")
        return None 

def to_pdf(df, obra_nome):
    try:
        pdf = FPDF(orientation='L', unit='mm', format='A4')
        pdf.add_page()
        logo_path = "Lavie.png"
        try:
            pdf.image(logo_path, x=10, y=3, w=50)
        except Exception as e:
            print(f"Aviso: Não foi possível carregar o logo '{logo_path}'. Erro: {e}")
        pdf.set_font('Arial', 'B', 16)
        
        titulo = f'Relatorio de Patrimonio - Obra: {obra_nome}'.encode('latin-1', 'replace').decode('latin-1')
        pdf.cell(0, 10, titulo, 0, 1, 'C')
        pdf.ln(10)

        pdf.set_font('Arial', 'B', 8)
        
        col_widths = {
            TOMBAMENTO_COL: 40, 
            NOME_COL: 60, 
            STATUS_COL: 30, 
            LOCAL_COL: 40, 
            RESPONSAVEL_COL: 40, 
            VALOR_COL: 25
        }
        cols_to_export = list(col_widths.keys())
        
        for col_name in cols_to_export:
            pdf.cell(col_widths[col_name], 7, col_name.replace("_", " ").title(), 1, 0, 'C')
        pdf.ln()

        pdf.set_font('Arial', '', 8)
        
        df_pdf = df[cols_to_export].fillna('') 
        
        for _, row in df_pdf.iterrows():
            for col_name in cols_to_export:
                text = str(row[col_name]).encode('latin-1', 'replace').decode('latin-1')
                pdf.cell(col_widths[col_name], 6, text, 1)
            pdf.ln()

        return bytes(pdf.output(dest='S'))
    
    except KeyError as e:
        st.error(f"Erro ao gerar PDF (KeyError): {e}")
        st.error("Isso geralmente acontece porque os dados estão vazios (devido a um erro anterior) ou as 'Constantes das Colunas' no código não batem com o Supabase.")
        return None
    except Exception as e:
        st.error(f"Erro inesperado ao gerar PDF: {e}")
        return None

def gerar_ficha_qr_code(row_series):
    """Gera um PDF com ficha técnica e QR Code para um único item."""
    try:
        pdf = FPDF(orientation='P', unit='mm', format='A4')
        pdf.add_page()
        
        pdf.set_fill_color(227, 112, 38)
        pdf.rect(0, 0, 210, 20, 'F')
        
        pdf.set_text_color(255, 255, 255)
        pdf.set_font('Helvetica', 'B', 16)
        pdf.text(10, 14, "Ficha de Identificação de Ativo - LAVIE")
        
        qr_data = f"ID: {row_series[ID_COL]}\nItem: {row_series[NOME_COL]}\nTombamento: {row_series[TOMBAMENTO_COL]}\nObra: {row_series[OBRA_COL]}"
        
        qr = qrcode.QRCode(box_size=10, border=4)
        qr.add_data(qr_data)
        qr.make(fit=True)
        img_qr = qr.make_image(fill_color="black", back_color="white")
        
        with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp_file:
            img_qr.save(tmp_file.name)
            qr_path = tmp_file.name

        pdf.set_text_color(0, 0, 0)
        pdf.set_y(30)
        pdf.set_font('Helvetica', 'B', 12)
        pdf.cell(0, 10, f"Produto: {str(row_series[NOME_COL]).upper()}", ln=True)
        
        pdf.set_font('Helvetica', '', 11)
        pdf.cell(0, 8, f"Tombamento: {row_series[TOMBAMENTO_COL]}", ln=True)
        pdf.cell(0, 8, f"Obra Atual: {row_series[OBRA_COL]}", ln=True)
        pdf.cell(0, 8, f"Responsável: {row_series[RESPONSAVEL_COL]}", ln=True)
        pdf.cell(0, 8, f"Status: {row_series[STATUS_COL]}", ln=True)
        
        pdf.ln(5)
        pdf.set_font('Helvetica', 'B', 10)
        pdf.cell(0, 8, "Especificações / Obs:", ln=True)
        pdf.set_font('Helvetica', '', 10)
        pdf.multi_cell(110, 6, f"{str(row_series[ESPEC_COL])}\n{str(row_series[OBS_COL])}")

        pdf.image(qr_path, x=130, y=30, w=60)
        
        pdf.set_y(100)
        pdf.set_font('Helvetica', 'I', 8)
        pdf.cell(0, 10, "Este código QR serve para identificação rápida e auditoria do ativo.", 0, 1, 'C')

        return bytes(pdf.output())
        
    except Exception as e:
        st.error(f"Erro ao gerar Ficha QR: {e}")
        return None

@st.dialog("Atualizar Status")
def modal_atualizar_status(id_equipamento, nome_equipamento, status_atual, responsavel_atual):
    st.write(f"Equipamento: **{nome_equipamento}**")
    opcoes_status = ["Ativa (Em Uso)", "Disponível", "Em Manutenção", "Descartado"]
    
    try:
        index_atual = opcoes_status.index(status_atual)
    except ValueError:
        index_atual = 0

    novo_status = st.selectbox(
        "Novo Status:",
        opcoes_status,
        index=index_atual
    )
    
    novo_responsavel = st.text_input("Responsável:", value=responsavel_atual)

    st.write("") 

    col_cancelar, col_salvar = st.columns([1, 1])

    with col_cancelar:
        if st.button("Cancelar", use_container_width=True):
            st.rerun()

    with col_salvar:
        if st.button("Salvar", type="primary", use_container_width=True):
            
            st.success("Atualizado!")
            time.sleep(0.5)
            st.rerun()
        
def tela_de_login():
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.image("Lavie.png", use_container_width=True) 
    st.title("Controle de Patrimônio")

    tab1, tab2 = st.tabs(["Acesso por Obra", "Acesso de Administrador"])

    with tab1:
        st.subheader("Login da Obra")
        try:
            _, lista_obras, _, _, _ = carregar_dados_app()
            
            if not lista_obras:
                st.info("Nenhuma obra cadastrada no sistema.")
                return

            codigos_obras = st.secrets.obra_codes
            obra_selecionada = st.selectbox("Selecione a Obra", options=lista_obras, index=None, placeholder="Escolha a obra...")
            if obra_selecionada:
                codigo_acesso = st.text_input("Código de Acesso", type="password", key="obra_password")
                if st.button("Entrar na Obra", type="primary"):
                    if codigo_acesso == codigos_obras.get(obra_selecionada):
                        st.session_state.logged_in = True
                        st.session_state.is_admin = False
                        st.session_state.selected_obra = obra_selecionada
                        st.rerun()
                    else:
                        st.error("Código de acesso incorreto.")
        except Exception as e:
            st.error(f"Não foi possível carregar a lista de obras. Erro: {e}")

    with tab2:
        st.subheader("Login de Administrador")
        admin_password = st.text_input("Senha do Administrador", type="password", key="admin_password")
        if st.button("Entrar como Administrador", type="primary"):
            if admin_password == st.secrets.admin.password:
                st.session_state.logged_in = True
                st.session_state.is_admin = True
                st.rerun()
            else:
                st.error("Senha de administrador incorreta.")
                
def pagina_dashboard(dados_da_obra, df_movimentacoes):
    st.header("Análise de Ativos", divider='rainbow')

    COR_PRINCIPAL = "#E37026"

    if dados_da_obra.empty:
        st.info("Nenhum dado de patrimônio disponível para exibir no dashboard.")
        return

    dados_com_idade = dados_da_obra.copy()
    idade_media_dias = None

    if not df_movimentacoes.empty:
        df_movimentacoes['data_hora'] = pd.to_datetime(df_movimentacoes['data_hora'])
        entradas = df_movimentacoes[df_movimentacoes['tipo_movimentacao'] == 'Entrada']
        
        if not entradas.empty:
            aquisicoes = entradas.groupby(TOMBAMENTO_COL)['data_hora'].min().reset_index()
            aquisicoes.rename(columns={'data_hora': 'data_aquisicao'}, inplace=True)
            
            dados_com_idade = pd.merge(
                dados_com_idade, 
                aquisicoes, 
                on=TOMBAMENTO_COL, 
                how='left'
            )
            
            if 'data_aquisicao' in dados_com_idade.columns:
                agora_utc = datetime.now(datetime.timezone.utc)
                dados_com_idade['data_aquisicao'] = pd.to_datetime(dados_com_idade['data_aquisicao'], utc=True)
                
                dados_com_idade['idade_dias'] = (agora_utc - dados_com_idade['data_aquisicao']).dt.days
                idade_media_dias = dados_com_idade['idade_dias'].mean()

    st.subheader("Visão Geral do Patrimônio")
    total_itens = dados_da_obra.shape[0]
    valor_total = dados_da_obra[VALOR_COL].sum()
    
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Total de Itens", f"{total_itens} un.")
    with col2:
        st.metric("Valor Total do Patrimônio", f"R$ {valor_total:,.2f}")
    with col3:
        if idade_media_dias is not None:
            st.metric("Idade Média dos Ativos", f"{idade_media_dias:,.0f} dias")
        else:
            st.metric("Idade Média dos Ativos", "N/A")
            
    st.write("---")

    st.subheader("Análise de Custo e Valor")
    col_v1, col_v2 = st.columns([1, 2])
    
    with col_v1:
        st.markdown("**Top 10 Ativos Mais Valiosos**")
        top_10_valiosos = dados_da_obra.sort_values(by=VALOR_COL, ascending=False).head(10)
        st.dataframe(
            top_10_valiosos[[NOME_COL, VALOR_COL, RESPONSAVEL_COL]], 
            use_container_width=True,
            column_config={
                VALOR_COL: st.column_config.NumberColumn(format="R$ %.2f")
            }
        )
    
    with col_v2:
        st.markdown("**Distribuição do Valor dos Ativos**")
        fig_hist_valor = px.histogram(
            dados_da_obra, 
            x=VALOR_COL, 
            nbins=50, 
            title="Histograma: Frequência de Itens por Faixa de Valor",
            text_auto=True
        )
        fig_hist_valor.update_traces(marker_color=COR_PRINCIPAL)
        fig_hist_valor.update_layout(yaxis_title="Contagem de Itens", xaxis_title="Valor (R$)")
        st.plotly_chart(fig_hist_valor, use_container_width=True)

    st.write("---")

    st.subheader("Análise de Aquisição e Operações ao Longo do Tempo")
    
    col_t1, col_t2 = st.columns(2)
    
    with col_t1:
        st.markdown("**Aquisição de Ativos ao Longo do Tempo**")
        if 'data_aquisicao' in dados_com_idade.columns and not dados_com_idade['data_aquisicao'].isnull().all():
            aquisicoes_no_tempo = dados_com_idade.set_index('data_aquisicao').resample('M')[VALOR_COL].sum().reset_index()
            aquisicoes_no_tempo = aquisicoes_no_tempo[aquisicoes_no_tempo[VALOR_COL] > 0]
            
            fig_aquisicao = px.line(
                aquisicoes_no_tempo, 
                x='data_aquisicao', 
                y=VALOR_COL, 
                title="Valor Adquirido por Mês",
                markers=True
            )
            fig_aquisicao.update_traces(line_color=COR_PRINCIPAL, marker_color=COR_PRINCIPAL)
            fig_aquisicao.update_layout(xaxis_title="Data da Aquisição", yaxis_title="Valor Adquirido (R$)")
            st.plotly_chart(fig_aquisicao, use_container_width=True)
        else:
            st.info("Não há dados de 'Entrada' suficientes na tabela de movimentações para gerar a análise de aquisição.")

    with col_t2:
        st.markdown("**Fluxo de Movimentações (Entrada vs. Saída)**")
        if not df_movimentacoes.empty:
            mov_no_tempo = df_movimentacoes.set_index('data_hora').groupby('tipo_movimentacao').resample('M').size().reset_index(name='contagem')
            
            color_map = {'Entrada': COR_PRINCIPAL, 'Saída': '#bec8c3'}
            
            fig_mov = px.line(
                mov_no_tempo,
                x='data_hora',
                y='contagem',
                color='tipo_movimentacao',
                color_discrete_map=color_map, 
                title="Movimentações por Mês",
                markers=True
            )
            fig_mov.update_layout(xaxis_title="Data da Movimentação", yaxis_title="Número de Movimentações")
            st.plotly_chart(fig_mov, use_container_width=True)
        else:
            st.info("Não há dados na tabela de movimentações.")
            
    st.write("---")
    
    st.subheader("Análise de Responsabilidade e Risco")
    
    col_r1, col_r2 = st.columns(2)
    
    with col_r1:
        st.markdown("**Valor Total (R$) por Responsável**")
        valor_por_resp = dados_da_obra.groupby(RESPONSAVEL_COL)[VALOR_COL].sum().sort_values(ascending=False).reset_index()
        fig_resp_val = px.bar(
            valor_por_resp,
            x=RESPONSAVEL_COL,
            y=VALOR_COL,
            title="Valor de Ativos por Responsável",
            text_auto='.2s'
        )
        fig_resp_val.update_traces(marker_color=COR_PRINCIPAL, textposition='outside')
        st.plotly_chart(fig_resp_val, use_container_width=True)

    with col_r2:
        st.markdown("**Análise de Status dos Ativos**")
        if not dados_com_idade.empty:
            status_counts = dados_com_idade[STATUS_COL].value_counts().reset_index()
            fig_status = px.pie(
                status_counts, 
                names=STATUS_COL, 
                values='count', 
                title="Distribuição de Itens por Status",
                color_discrete_sequence=px.colors.sequential.Oranges_r 
            )
            st.plotly_chart(fig_status, use_container_width=True)
        else:
            st.info("Não há dados de status para analisar.")

def pagina_cadastrar_item(is_admin, lista_status, lista_obras_app, existing_data):
    st.header("Novo Cadastro", divider='rainbow')
    
    tab_patrimonio, tab_locacao = st.tabs(["Patrimônio", "Locação"])

    with tab_patrimonio:
        st.markdown("### Registrar Novo Ativo")
        
        obra_para_cadastro = None
        if is_admin:
            obra_para_cadastro = st.selectbox("Obra de Destino", options=lista_obras_app, index=None, key="patr_obra_sel")
        else:
            obra_para_cadastro = st.session_state.selected_obra

        if not obra_para_cadastro:
            st.info("Selecione uma obra acima para liberar o formulário.")
        else:
            with st.form("cadastro_patrimonio_form", clear_on_submit=True):
                p1_c1, p1_c2, p1_c3 = st.columns([3, 1.5, 1.5])
                with p1_c1:
                    nome_produto = st.text_input("Nome do Produto/Ativo")
                with p1_c2:
                    num_tombamento_manual = st.text_input("Tombamento (Opcional)")
                with p1_c3:
                    status_selecionado = st.selectbox("Status Inicial", options=lista_status, index=0)

                p2_c1, p2_c2, p2_c3, p2_c4 = st.columns([2, 2, 1.5, 1.5])
                with p2_c1:
                    local_uso = st.text_input("Local de Uso (Ex: Almoxarifado)")
                with p2_c2:
                    responsavel = st.text_input("Responsável Pelo Ativo")
                with p2_c3:
                    num_nota_fiscal = st.text_input("N° Nota Fiscal")
                with p2_c4:
                    valor_produto = st.number_input("Valor (R$)", min_value=0.0, step=100.00, format="%.2f")

                p3_c1, p3_c2 = st.columns(2)
                with p3_c1:
                    especificacoes = st.text_area("Especificações Técnicas", height=100)
                with p3_c2:
                    observacoes = st.text_area("Observações Gerais", height=100)
            
                st.write("---")
                uploaded_pdf = st.file_uploader("Anexar PDF da Nota Fiscal", type="pdf")
                
                submitted = st.form_submit_button("Cadastrar Patrimônio", type="primary", use_container_width=True)

                if submitted:
                    if not (nome_produto and num_nota_fiscal and local_uso and responsavel):
                        st.error("⚠️ Preencha os campos obrigatórios: Nome, NF, Local e Responsável.")
                    else:
                        link_nota_fiscal = ""
                        num_final_envio = num_tombamento_manual.strip() if num_tombamento_manual else None

                        if uploaded_pdf:
                            file_name = f"NF_{obra_para_cadastro}_{datetime.now().strftime('%H%M%S')}.pdf"
                            link_nota_fiscal = upload_to_supabase_storage(uploaded_pdf.getvalue(), file_name)

                        novo_item_dict = {
                            OBRA_COL: obra_para_cadastro,
                            TOMBAMENTO_COL: num_final_envio,
                            NOME_COL: nome_produto,
                            ESPEC_COL: especificacoes,
                            OBS_COL: observacoes,
                            LOCAL_COL: local_uso,
                            RESPONSAVEL_COL: responsavel,
                            NF_NUM_COL: num_nota_fiscal,
                            NF_LINK_COL: link_nota_fiscal,
                            VALOR_COL: valor_produto,
                            STATUS_COL: status_selecionado
                        }
                        
                        try:
                            conn.table("patrimonio").insert(novo_item_dict).execute()
                            st.success(f"Patrimônio '{nome_produto}' cadastrado com sucesso!")
                            st.cache_data.clear() 
                            st.rerun()
                        except Exception as e:
                            st.error(f"Erro ao salvar: {e}")

    with tab_locacao:
        st.markdown("### Registrar Nova Locação")
        
        with st.form("cadastro_locacao_form", clear_on_submit=True):
            l1_c1, l1_c2, l1_c3, l1_c4 = st.columns([2, 2, 2, 1])
            with l1_c1:
                loc_equipamento = st.text_input("Equipamento")
            with l1_c2:
                if is_admin:
                    loc_obra = st.selectbox("Obra Destino", options=lista_obras_app, key="loc_obra_select")
                else:
                    loc_obra = st.text_input("Obra Destino", value=st.session_state.selected_obra, disabled=True)
                    if not loc_obra: loc_obra = st.session_state.selected_obra
            with l1_c3:
                loc_responsavel = st.text_input("Responsável (Rastreio)")
            with l1_c4:
                loc_qtd = st.number_input("Quantidade", min_value=1, value=1, step=1)

            l2_c1, l2_c2, l2_c3, l2_c4 = st.columns([1, 1.5, 1.5, 2])
            with l2_c1:
                loc_unidade = st.text_input("Unidade (Ex: Mês)")
            with l2_c2:
                loc_valor = st.number_input("Valor Unitário/Mensal (R$)", min_value=0.0, format="%.2f")
            with l2_c3:
                loc_contrato = st.text_input("Contrato/PC (Sienge)")
            with l2_c4:
                loc_status = st.selectbox("Status Inicial", [ "Ativo", "Manutenção", "Devolvido"])

            l3_c1, l3_c2 = st.columns(2)
            with l3_c1:
                loc_inicio = st.date_input("Data de Início da Cobrança", value=None)
            with l3_c2:
                loc_fim = st.date_input("Previsão Fim da Locação", value=None)

            st.write("")
            submitted_loc = st.form_submit_button("Adicionar Locação", type="primary", use_container_width=True)

            if submitted_loc:
                if not (loc_equipamento and loc_obra):
                    st.error("Campos 'Equipamento' e 'Obra' são obrigatórios.")
                else:
                    nova_locacao = {
                        "equipamento": loc_equipamento,
                        "obra_destino": loc_obra if is_admin else st.session_state.selected_obra,
                        "responsavel": loc_responsavel,
                        "quantidade": loc_qtd,
                        "unidade": loc_unidade,
                        "valor_mensal": loc_valor,
                        "contrato_sienge": loc_contrato,
                        "status": loc_status,
                        "data_inicio": loc_inicio.isoformat() if loc_inicio else None,
                        "data_previsao_fim": loc_fim.isoformat() if loc_fim else None
                    }
                    try:
                        conn.table("locacoes").insert(nova_locacao).execute()
                        st.success(f"Locação de '{loc_equipamento}' registrada!")
                        st.cache_data.clear()
                        st.rerun()
                    except Exception as e:
                        st.error(f"Erro ao salvar locação: {e}")
                        
def pagina_itens_cadastrados(is_admin, dados_patrimonio, dados_locacoes, lista_status):
    st.header("Consulta e Relatórios", divider='rainbow')
    st.markdown("""
    <style>
    /* Força o botão "secondary" a ser AZUL (Estilo 'Atualizar Status') */
    button[kind="secondary"] {
        background-color: transparent !important;
        color: white !important;
        border: none !important;
        font-weight: 500 !important;
    }
    button[kind="secondary"]:hover {
        background-color: transparent !important;
        color: white !important;
    }

    div[data-testid="stVerticalBlockBorderWrapper"] {
        background-color: #1E1E1E; /* Cor de fundo do card */
        border: 1px solid #333;
        border-radius: 10px;
        padding: 15px;
        margin-bottom: 15px;
        box-shadow: 0 4px 6px rgba(0,0,0,0.3);
    }
    </style>
    """, unsafe_allow_html=True)
    
    tab_vis_patrimonio, tab_vis_locacao = st.tabs(["Patrimônio", "Locações Ativas"])

    with tab_vis_patrimonio:
        if dados_patrimonio.empty:
            st.info("Nenhum patrimônio cadastrado.")
        else:
            col_f1, col_f2 = st.columns(2)
            with col_f1:
                search_term = st.text_input("Buscar Patrimônio", key="search_patrimonio", placeholder="Nome, Tombamento ou Responsável...")
            
            dados_filt = dados_patrimonio.copy()
            if search_term:
                dados_filt = dados_filt[
                    dados_filt[NOME_COL].str.contains(search_term, case=False, na=False) |
                    dados_filt[TOMBAMENTO_COL].astype(str).str.contains(search_term, case=False, na=False) |
                    dados_filt[RESPONSAVEL_COL].str.contains(search_term, case=False, na=False)
                ]

            total_valor_patr = dados_filt[VALOR_COL].sum()
            qtd_patr = dados_filt.shape[0]
            
            st.markdown(f"""
            <div style="background-color: transparent; background-image: linear-gradient(160deg, #1e1e1f 0%, #0a0a0c 100%); border: 1px solid rgba(255, 255, 255, 0.1); padding: 20px; border-radius: 10px; margin-botton:20px;">
                <h4 style="margin:0; color: #E37026;">Resumo Patrimonial</h4>
                <div style="display:flex; justify-content:space-between; align-items:center; margin-top:5px;">
                    <span><b>{qtd_patr}</b> itens encontrados</span>
                    <span style="font-size: 0.75rem; color: #FFFFFF; font-weight: 700; letter-spacing: 1.5px;"><b>VALOR TOTAL: R$ {total_valor_patr:,.2f}</b></span>
                </div>
            </div>
            """, unsafe_allow_html=True)

            st.markdown("")

            for index, row in dados_filt.iterrows():
                with st.container(border=False):
                    st_txt = str(row[STATUS_COL]).strip().upper()
                    
                    if st_txt == "ATIVO":
                        cor_status = "#35BE53" 
                    elif st_txt in ["MANUTENÇÃO", "EMPRÉSTIMO"]:
                        cor_status = "#ffc107" 
                    else:
                        cor_status = "#dc3545" 
                    bg_status = f"{cor_status}22" 
                    valor_fmt = f"R$ {row[VALOR_COL]:,.2f}"
                    nome_safe = str(row[NOME_COL]).replace('"', '&quot;')
                    espec_safe = str(row[ESPEC_COL])[:100] + "..."

                    html_content = f"""
                    <div style="margin-bottom: 10px;">
                        <div style="display:flex; justify-content:space-between; align-items:start;">
                            <div>
                                <h3 style="margin:0; color: white; font-size: 1.3em;">{nome_safe}</h3>
                                <div style="color: #E37026; font-weight: bold; font-size: 0.9em;">TOMBAMENTO: {row[TOMBAMENTO_COL]}</div>
                            </div>
                            <span style="background-color: {bg_status}; color: {cor_status}; padding: 4px 12px; border-radius: 10px; font-size: 0.75em; border: 1px solid {cor_status}; font-weight: bold;">{st_txt}</span>
                        </div>
                        <div style="margin-top: 15px; display:flex; flex-wrap: wrap; gap: 20px; color: #CCC; font-size: 0.9em;">
                            <div style="min-width: 120px;"><b style="color: #888; display:block;">OBRA</b>{row[OBRA_COL]}</div>
                            <div style="min-width: 120px;"><b style="color: #888; display:block;">LOCAL</b>{row[LOCAL_COL]}</div>
                            <div style="min-width: 120px;"><b style="color: #888; display:block;">RESPONSÁVEL</b>{row[RESPONSAVEL_COL]}</div>
                            <div><b style="color: #888; display:block;">VALOR</b><span style="color: #4cd137;">{valor_fmt}</span></div>
                        </div>
                        <div style="margin-top: 10px; font-size: 0.85em; color: #888; font-style: italic;">
                            {espec_safe}
                        </div>
                        <hr style="border-top: 1px solid #333; margin: 15px 0 10px 0;">
                    </div>
                    """
                    st.markdown(html_content, unsafe_allow_html=True)
                    
                    c_vazio, c_btn_nf, c_btn_qr = st.columns([4, 1.5, 1.5])
                    
                    with c_btn_nf:
                        if row[NF_LINK_COL]:
                            st.link_button("Nota Fiscal", row[NF_LINK_COL], type="primary", use_container_width=True)
                        else:
                            st.button("Sem Nota", disabled=True, key=f"btn_nf_{row[ID_COL]}", type="secondary", use_container_width=True)

                    with c_btn_qr:
                        if st.button("Etiqueta QR", key=f"btn_qr_{row[ID_COL]}", type="primary", use_container_width=True):
                            pdf_bytes = gerar_ficha_qr_code(row)
                            if pdf_bytes:
                                b64 = base64.b64encode(pdf_bytes).decode()
                                href = f'<a href="data:application/pdf;base64,{b64}" download="Etiqueta_{row[TOMBAMENTO_COL]}.pdf" style="display:none;" id="dl_link_{row[ID_COL]}">Download</a><script>document.getElementById("dl_link_{row[ID_COL]}").click();</script>'
                                st.markdown(f'<a href="data:application/pdf;base64,{b64}" download="Etiqueta_{row[TOMBAMENTO_COL]}.pdf" style="color:#E37026; text-decoration:none; font-weight:bold; display:block; text-align:center;">⬇️ Baixar PDF</a>', unsafe_allow_html=True)

    with tab_vis_locacao:
        if dados_locacoes.empty:
            st.info("Nenhuma locação registrada.")
            return

        col_lf1, col_lf2 = st.columns([1, 2])
        with col_lf1:
            obras_loc_disp = sorted(list(dados_locacoes["obra_destino"].unique()))
            filtro_obra_loc = st.selectbox("Filtrar por Obra", ["Todas"] + obras_loc_disp)
        with col_lf2:
            busca_loc = st.text_input("Busca Geral", key="search_loc", placeholder="Equipamento, contrato...")

        df_l = dados_locacoes.copy()
        if filtro_obra_loc != "Todas":
            df_l = df_l[df_l["obra_destino"] == filtro_obra_loc]
        if busca_loc:
            df_l = df_l[df_l["equipamento"].str.contains(busca_loc, case=False, na=False) | df_l["contrato_sienge"].str.contains(busca_loc, case=False, na=False)]

        total_mensal = df_l["valor_mensal"].sum()
        qtd_equip = df_l.shape[0]
        
        st.markdown(f"""
        <div style="background-color: rgba(227, 112, 38, 0.15); padding: 15px; border-radius: 10px; border: 1px solid #E37026; margin-bottom: 20px;">
            <div style="display:flex; justify-content:space-between; align-items:center;">
                <h4 style="margin:0; color: #E37026;">{filtro_obra_loc if filtro_obra_loc != 'Todas' else 'Geral'}</h4>
                <div style="text-align:right;">
                    <div style="font-size: 0.9em; color: #ccc;">VALOR TOTAL MENSAL ESTIMADO</div>
                    <div style="font-size: 1.4em; font-weight:bold;">R$ {total_mensal:,.2f}</div>
                </div>
            </div>
             <div style="font-size: 0.9em; color: #aaa; margin-top:5px;">{qtd_equip} equipamento(s) locado(s)</div>
        </div>""", unsafe_allow_html=True)

        for index, row in df_l.iterrows():
            with st.container(border=False):
                d_inicio = pd.to_datetime(row['data_inicio']).strftime('%d/%m/%Y') if pd.notnull(row['data_inicio']) else '-'
                d_fim = pd.to_datetime(row['data_previsao_fim']).strftime('%d/%m/%Y') if pd.notnull(row['data_previsao_fim']) else '-'
                valor_loc_fmt = f"R$ {row['valor_mensal']:,.2f}"
                equip_safe = str(row['equipamento']).replace('"', '&quot;')
                
                st_loc = str(row['status'])
                cor_loc = "#dc3545" 
                if "Ativo" in st_loc or "Manutenção" in st_loc: cor_loc = "#0d6efd"
                bg_loc = f"{cor_loc}22"

                html_loc = f"""
                <div style="margin-bottom: 10px;">
                    <div style="display:flex; justify-content:space-between; align-items:start;">
                        <div>
                            <h3 style="margin:0; color: white; font-size: 1.3em;">{equip_safe}</h3>
                            <span style="color: #888; font-size: 0.9em;">{row['contrato_sienge']}</span>
                        </div>
                        <span style="background-color: {bg_loc}; color: {cor_loc}; padding: 4px 10px; border-radius: 4px; font-size: 0.8em; border: 1px solid {cor_loc};">{st_loc}</span>
                    </div>
                    <div style="margin-top: 15px; display:flex; flex-wrap:wrap; gap: 20px; color: #CCC; font-size: 0.9em;">
                        <div style="min-width: 140px;"><b style="color: #888; display:block;">OBRA</b>{row['obra_destino']}</div>
                        <div style="min-width: 80px;"><b style="color: #888; display:block;">QTD</b>{row['quantidade']}</div>
                        <div style="min-width: 140px;"><b style="color: #888; display:block;">RESPONSÁVEL</b>{row['responsavel']}</div>
                        <div><b style="color: #888; display:block;">VALOR</b><span style="color: #4cd137;">{valor_loc_fmt}</span></div>
                    </div>
                    <div style="margin-top: 10px; font-size: 0.85em; color: #aaa; display:flex; gap: 20px;">
                        <span>Início: {d_inicio}</span><span>Prev. Fim: {d_fim}</span>
                    </div>
                    <hr style="border-top: 1px solid #333; margin: 15px 0 10px 0;">
                </div>
                """
                st.markdown(html_loc, unsafe_allow_html=True)
                
                c_vaz, c_btn1, c_btn2 = st.columns([5, 2.5, 2.5])
                
                with c_btn1:
                    if st.button("Atualizar Status", key="btn_update_1", type="primary"):
                        modal_atualizar_status(
                            id_equipamento=row['id'],
                            nome_equipamento=row['equip_safe'],
                            status_atual=row['status'],
                            responsavel_atual=row['responsavel']
                        )
                            
                
                with c_btn2:
                    if st.button("Excluir Registro", key=f"btn_del_{row['id']}", type="primary", use_container_width=True):
                        conn.table("locacoes").delete().eq("id", row['id']).execute()
                        st.cache_data.clear()
                        st.rerun()
                
                st.markdown("---")
                
def pagina_gerenciar_itens(dados_da_obra, existing_data_full, df_movimentacoes, lista_status):
    st.header("Gerenciar Itens Cadastrados", divider='rainbow')

    if dados_da_obra.empty:
        st.info("Nenhum item cadastrado para a obra selecionada ainda.")
        return

    dados_filtrados_gerenciar = dados_da_obra.copy()
    
    with st.expander("Filtros", expanded=True):
        col_f1, col_f2 = st.columns(2)
        with col_f1:
            status_unicos_ger = ["Todos"] + sorted(list(dados_da_obra[STATUS_COL].unique()))
            filtro_status_ger = st.selectbox("Filtrar por Status", status_unicos_ger, key="filter_status_ger")
            if filtro_status_ger != "Todos":
                dados_filtrados_gerenciar = dados_filtrados_gerenciar[dados_filtrados_gerenciar[STATUS_COL] == filtro_status_ger]
        
        with col_f2:
            search_term_ger = st.text_input("Buscar por Nome, Tombamento ou Responsável", key="filter_search_ger")
            if search_term_ger:
                dados_filtrados_gerenciar = dados_filtrados_gerenciar[
                    dados_filtrados_gerenciar[NOME_COL].str.contains(search_term_ger, case=False, na=False) |
                    dados_filtrados_gerenciar[TOMBAMENTO_COL].astype(str).str.contains(search_term_ger, case=False, na=False) |
                    dados_filtrados_gerenciar[RESPONSAVEL_COL].str.contains(search_term_ger, case=False, na=False)
                ]
    
    st.dataframe(dados_filtrados_gerenciar, use_container_width=True, hide_index=True, height=500)
    st.write("---")
 
    lista_itens = [f"{row[TOMBAMENTO_COL]} - {row[NOME_COL]} (ID: {row[ID_COL]})" for _, row in dados_filtrados_gerenciar.iterrows()]
    item_selecionado_gerenciar = st.selectbox("Selecione um item para Gerenciar", options=lista_itens, index=None, placeholder="Escolha um item...")

    if item_selecionado_gerenciar:
        item_id_selecionado = int(item_selecionado_gerenciar.split("(ID: ")[1].replace(")", ""))
        
        item_data_series_list = dados_filtrados_gerenciar[dados_filtrados_gerenciar[ID_COL] == item_id_selecionado]
        
        if item_data_series_list.empty:
            st.error("Item não encontrado. Por favor, atualize a página.")
            return
            
        item_data_series = item_data_series_list.iloc[0]
        tombamento_selecionado = item_data_series[TOMBAMENTO_COL]
        obra_do_item = item_data_series[OBRA_COL]
        
        if not st.session_state.get('confirm_delete'):
            col_mov, col_edit, col_delete = st.columns(3)
            
            if col_mov.button("Registrar Entrada/Saída",type="primary", use_container_width=True):
                st.session_state.movement_item_id = item_id_selecionado
                st.session_state.edit_item_id = None
                st.session_state.confirm_delete = False
                st.rerun()

            if col_edit.button("Editar Item", type="primary", use_container_width=True):
                st.session_state.edit_item_id = item_id_selecionado
                st.session_state.movement_item_id = None
                st.session_state.confirm_delete = False
                st.rerun()
        else:
            col_delete = st.container() 

        if col_delete.button("Remover Item", type="primary", use_container_width=True):
            st.session_state.edit_item_id = item_id_selecionado
            st.session_state.confirm_delete = True
            st.session_state.movement_item_id = None
            st.rerun()

        if st.session_state.confirm_delete and st.session_state.edit_item_id == item_id_selecionado:
            st.warning(f"**Atenção!** Tem certeza que deseja remover permanentemente o item **{tombamento_selecionado}** da obra **{obra_do_item}**?")
            c1_del, c2_del = st.columns(2)
            
            if c1_del.button("Sim, tenho certeza e quero remover", use_container_width=True, type="primary"):
                try:
                    conn.table("patrimonio").delete().eq(ID_COL, item_id_selecionado).execute()
                    st.success(f"Item {tombamento_selecionado} da obra {obra_do_item} removido!")
                    st.session_state.confirm_delete = False
                    st.session_state.edit_item_id = None
                    st.cache_data.clear()
                    st.rerun()
                except Exception as e:
                    st.error(f"Erro ao remover item: {e}")

            if c2_del.button("Cancelar", use_container_width=True):
                st.session_state.confirm_delete = False
                st.session_state.edit_item_id = None
                st.rerun()
        
        if st.session_state.movement_item_id == item_id_selecionado:
            with st.form("movement_form"):
                st.subheader(f"Registrar Movimentação para: {item_selecionado_gerenciar}")
                tipo_mov = st.radio("Tipo de Movimentação", ["Entrada", "Saída"], horizontal=True)
                responsavel_mov = st.text_input("Responsável pela Movimentação*")
                obs_mov = st.text_area("Observações da Movimentação")
                submitted_mov = st.form_submit_button("Registrar Movimentação")
            
                if submitted_mov:
                    if not responsavel_mov:
                        st.warning("O campo 'Responsável pela Movimentação' é obrigatório.")
                    else:
                        nova_movimentacao = {
                            OBRA_COL: obra_do_item,
                            TOMBAMENTO_COL: tombamento_selecionado,
                            "tipo_movimentacao": tipo_mov,
                            "data_hora": datetime.now().isoformat(),
                            "responsavel_movimentacao": responsavel_mov,
                            OBS_COL: obs_mov
                        }
                        novo_status = "Disponível" if tipo_mov == "Entrada" else "Em Uso Externo"
                        
                        try:
                            conn.table("movimentacoes").insert(nova_movimentacao).execute()
                            conn.table("patrimonio").update({STATUS_COL: novo_status}).eq(ID_COL, item_id_selecionado).execute()
                            st.success("Movimentação registrada com sucesso!")
                            st.session_state.movement_item_id = None
                            st.cache_data.clear()
                            st.rerun()
                        except Exception as e:
                            st.error(f"Erro ao registrar movimentação: {e}")
                        
        if st.session_state.edit_item_id == item_id_selecionado and not st.session_state.confirm_delete:
            with st.form("edit_form"):
                st.subheader(f"Editando Item: {tombamento_selecionado} (Obra: {obra_do_item})")
                
                tomb_edit_novo = st.text_input(f"{TOMBAMENTO_COL}", value=item_data_series.get(TOMBAMENTO_COL, ""))
                status_edit = st.selectbox(STATUS_COL, options=lista_status, index=lista_status.index(item_data_series.get(STATUS_COL)) if item_data_series.get(STATUS_COL) in lista_status else 0)
                nome_edit = st.text_input(NOME_COL, value=item_data_series.get(NOME_COL, ""))
                num_nota_fiscal_edit = st.text_input(f"{NF_NUM_COL}*", value=item_data_series.get(NF_NUM_COL, ""))
                especificacoes_edit = st.text_area(ESPEC_COL, value=item_data_series.get(ESPEC_COL, ""))
                observacoes_edit = st.text_area(OBS_COL, value=item_data_series.get(OBS_COL, ""))
                local_edit = st.text_input(LOCAL_COL, value=item_data_series.get(LOCAL_COL, ""))
                responsavel_edit = st.text_input(RESPONSAVEL_COL, value=item_data_series.get(RESPONSAVEL_COL, ""))
                valor_edit = st.number_input(f"{VALOR_COL} (R$)", min_value=0.0, format="%.2f", value=float(item_data_series.get(VALOR_COL, 0)))
                
                submitted_edit = st.form_submit_button("Salvar Alterações")
                
                if submitted_edit:
                    if not num_nota_fiscal_edit or not tomb_edit_novo:
                        st.warning(f"Os campos '{TOMBAMENTO_COL}' e '{NF_NUM_COL}*' são obrigatórios.")
                    else: 
                        edit_input_limpo = tomb_edit_novo.strip()
                        
                        condicao_outro_item = (existing_data_full[OBRA_COL] == obra_do_item) & \
                                              (existing_data_full[TOMBAMENTO_COL].astype(str).str.strip() == edit_input_limpo) & \
                                              (existing_data_full[ID_COL] != item_id_selecionado)
                        
                        if not existing_data_full[condicao_outro_item].empty:
                            st.error(f"Erro: O N° de Tombamento '{edit_input_limpo}' já existe para outro item nesta obra.")
                        else:
                            update_dict = {
                                TOMBAMENTO_COL: edit_input_limpo,
                                STATUS_COL: status_edit,
                                NOME_COL: nome_edit,
                                NF_NUM_COL: num_nota_fiscal_edit,
                                ESPEC_COL: especificacoes_edit,
                                OBS_COL: observacoes_edit,
                                LOCAL_COL: local_edit,
                                RESPONSAVEL_COL: responsavel_edit,
                                VALOR_COL: valor_edit
                            } 
                            
                            try:
                                conn.table("patrimonio").update(update_dict).eq(ID_COL, item_id_selecionado).execute()
                                st.success(f"Item {edit_input_limpo} atualizado com sucesso!")
                                st.session_state.edit_item_id = None
                                st.cache_data.clear()
                                st.rerun()
                            except Exception as e:
                                st.error(f"Erro ao atualizar item: {e}")
            
        st.write("---")
        st.subheader(f"Histórico de Movimentações do Item: {tombamento_selecionado}")
        historico_item = df_movimentacoes[
            (df_movimentacoes[OBRA_COL] == obra_do_item) &
            (df_movimentacoes[TOMBAMENTO_COL].astype(str) == str(tombamento_selecionado))
        ].sort_values(by="data_hora", ascending=False)
    
        if not historico_item.empty:
            st.dataframe(historico_item, hide_index=True, use_container_width=True, column_config={ID_COL: None})
        else:
            st.info("Nenhuma movimentação registrada para este item.")

def app_principal():
    is_admin = st.session_state.is_admin
    obra_selecionada_sidebar = None 

    lista_status, lista_obras_app, existing_data_full, df_movimentacoes, df_locacoes = carregar_dados_app()
    
    with st.sidebar:
        st.image("Lavie.png", use_container_width=True)
        st.header("Navegação")
        if is_admin:
            st.info("Logado como **Administrador**.")
        else:
            st.info(f"Obra: **{st.session_state.selected_obra}**")

        menu_options = ["Cadastrar Item", "Consulta Geral", "Gerenciar Itens", "Dashboard"]
        icons = ["plus-circle-fill", "card-list", "pencil-square", "bar-chart-fill"]
        
        selected_page = option_menu(
            menu_title=None,
            options=menu_options,
            icons=icons,
            menu_icon="cast",
            default_index=0,
            styles={ 
                "container": {"padding": "5px !important", "background-color": "transparent"},
                "icon": {"font-size": "18px"}, 
                "nav-link": {"font-size": "16px", "text-align": "left", "margin":"0px"},
                "nav-link-selected": {"background-color": "#E37026"}, 
            }
        )

        if is_admin:
            st.write("---")
            obras_disponiveis = ["Todas"] + lista_obras_app 
            obra_selecionada_sidebar = st.selectbox("Filtrar Visão por Obra", obras_disponiveis)
            
        nome_da_obra_para_relatorio = "" 
    
    if is_admin:
        nome_da_obra_para_relatorio = obra_selecionada_sidebar
        if obra_selecionada_sidebar == "Todas":
            dados_patrimonio = existing_data_full
            dados_locacoes_filt = df_locacoes
        else:
            dados_patrimonio = existing_data_full[existing_data_full[OBRA_COL] == obra_selecionada_sidebar].copy()
            dados_locacoes_filt = df_locacoes[df_locacoes["obra_destino"] == obra_selecionada_sidebar].copy()
            
    else: 
        obra_logada = st.session_state.selected_obra
        nome_da_obra_para_relatorio = obra_logada
        dados_patrimonio = existing_data_full[existing_data_full[OBRA_COL] == obra_logada].copy()
        dados_locacoes_filt = df_locacoes[df_locacoes["obra_destino"] == obra_logada].copy()

    with st.sidebar:
        st.write("---")
        st.header("Exportação Rápida")
        if st.button("Sair / Trocar Obra", type="primary", use_container_width=True):
            for key in st.session_state.keys():
                del st.session_state[key]
            st.cache_data.clear()
            st.rerun()

    if selected_page == "Dashboard":
        pagina_dashboard(dados_patrimonio, df_movimentacoes)
        
    elif selected_page == "Cadastrar Item":
        pagina_cadastrar_item(is_admin, lista_status, lista_obras_app, dados_patrimonio)
        
    elif selected_page == "Consulta Geral":
        pagina_itens_cadastrados(is_admin, dados_patrimonio, dados_locacoes_filt, lista_status)
        
    elif selected_page == "Gerenciar Itens":
        pagina_gerenciar_itens(dados_patrimonio, existing_data_full, df_movimentacoes, lista_status)

if not st.session_state.logged_in:
    tela_de_login()
else:
    app_principal()
